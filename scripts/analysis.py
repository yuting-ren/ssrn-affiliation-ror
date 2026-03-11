import pandas as pd
import matplotlib.pyplot as plt
import datetime
import matplotlib.dates as mdates
from scipy.stats import ttest_ind
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side

def analysis(output_computations, output_analysis_1, output_analysis_2, interest_name, interest_year, interest_month, interest_day):

    print('Computing analysis (step 6/6)')

    db = pd.read_csv(output_computations, sep=';', low_memory=False)
    db['approved_date'] = pd.to_datetime(db['approved_date'], dayfirst=True)

    metrics = [
        'flesh_reading_ease',
        'fk_grade_level',
        'gunning_fog',
        'smog',
        'dale_chall',
        'automated_readility',
        'avg_length_words',
        'prop_more_15words',
        'ttr',
        'page_count',
        'cle'
    ]

    colors = {
        'fk_grade_level':'tab:cyan',
        'gunning_fog': 'tab:blue',
        'smog': 'tab:orange',
        'automated_readility': 'tab:green',
        'flesh_reading_ease': 'tab:red',
        'dale_chall': 'tab:purple',
        'ttr': 'tab:brown',
        'avg_length_words': 'tab:pink',
        'prop_more_15words': 'tab:gray',
        'page_count': 'tab:olive',
        'cle': 'yellow'
    }

    daily_avg = {}
    monthly_avg = {}

    ticks_date = pd.date_range(start=db['approved_date'].min(), end=db['approved_date'].max(), freq='12ME')

    # If the dataset is small (under 2 years), the frequence of the ticks is every month instead of every 3 months
    if len(ticks_date) >= 2:
        frequence = '3ME'
    else:
        frequence = 'ME'

    # Computing the average to plot them and to smooth them out 
    for metric in metrics:
        daily_avg[metric] = db.groupby('approved_date')[metric].mean().reset_index()
        monthly_avg[metric] = daily_avg[metric].groupby(pd.Grouper(key='approved_date', freq=frequence))[metric].mean().reset_index()

    # Ploting all of the metrics and putting the interest date as a reference
    target_date = datetime.datetime(interest_year, interest_month, interest_day)
    cutoff_date = pd.Timestamp(target_date)
    ticks = list(mdates.date2num(ticks_date)) + [mdates.date2num(target_date)]
    labels = [pd.to_datetime(mdates.num2date(t)).strftime('%Y-%m') for t in ticks]
    labels[-1] = target_date.strftime('%Y-%m')

    plt.figure(figsize=(16, 8))
    for metric in metrics:
        plt.plot(monthly_avg[metric]['approved_date'], monthly_avg[metric][metric], label=metric, color=colors[metric])
    plt.title('Monthly average of all metrics')
    plt.axvline(x=target_date, color='red', linestyle='--', linewidth=1, label=interest_name)
    plt.xticks(ticks, labels, rotation=45)
    plt.xlabel('Date')
    plt.ylabel('Average value')
    plt.legend(ncol=2)
    plt.grid(True)
    plt.tight_layout()
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
    plt.savefig('outputs/graphs/all_papers/monthly_average_all_metrics.png')
    plt.show()

    # Plot each metric individually
    for metric in metrics:
        plt.figure(figsize=(16, 8))
        plt.plot(monthly_avg[metric]['approved_date'], monthly_avg[metric][metric], label=metric, color=colors[metric])
        plt.axvline(x=target_date, color='red', linestyle='--', linewidth=1, label=interest_name)
        plt.xticks(ticks, labels, rotation=45)
        plt.title(f'Monthly average of {metric.replace("_", " ")}')
        plt.xlabel('Date')
        plt.ylabel('Average value')
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        plt.savefig(f'outputs/graphs/all_papers/monthly_average_{metric}.png')
        plt.close()

    results = []

    # Computing t-test between before and after the event 
    for metric in metrics:
        before = db.loc[db['approved_date'] <= cutoff_date, metric].dropna()
        after = db.loc[db['approved_date'] > cutoff_date, metric].dropna()

        if len(before) > 1 and len(after) > 1:
            t_stat, p_val = ttest_ind(before, after, equal_var=False)
            results.append({'metric': metric, 't_stat': t_stat, 'p_value': p_val, 'mean_before': before.mean(), 'mean_after': after.mean()})
        else:
            results.append({'metric': metric, 't_stat': None, 'p_value': None, 'mean_before': before.mean(), 'mean_after': after.mean()})

    results_df = pd.DataFrame(results)
    results_df['significant'] = results_df['p_value'] < 0.05
    results_df['significant'] = results_df['significant'].map({True: 'True', False: 'False'})

    # Descriptive statistics + t-test
    describe_list = []

    for metric in metrics:
        desc = db[metric].describe()
        desc_dict = desc.to_dict()
        desc_dict['metric'] = metric
        describe_list.append(desc_dict)

    describe_df = pd.DataFrame(describe_list)

    final_df = pd.merge(describe_df,results_df, on='metric', how='left')
    final_df = final_df[['metric','count' ,'mean', 'std', 'min', '25%', '50%', '75%', 'max', 
                     't_stat', 'p_value', 'mean_before', 'mean_after', 'significant']]
    final_df = final_df.round(4)

    final_df.to_excel(output_analysis_1, index=False, engine='openpyxl')

    # Formatting of the Excel Output
    wb_2 = openpyxl.load_workbook(output_analysis_1)
    ws_2 = wb_2.active
    light_blue_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
    light_gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

    for cell in ws_2[1]:
        cell.fill = light_blue_fill
        cell.font = bold_font

    for row in ws_2.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.fill = light_blue_fill
    
    max_row = ws_2.max_row
    max_col = ws_2.max_column

    for row in ws_2.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.row == 1:
                cell.fill = light_blue_fill
                cell.font = bold_font
            elif cell.column == 1:
                cell.fill = light_blue_fill
                cell.font = bold_font 
            else:
                cell.fill = light_gray_fill
            cell.border = thin_border

    for col_idx, col in enumerate(ws_2.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value is not None:
                text = str(cell.value)
                max_length = max(max_length, len(text))
        adjusted_width = max(max_length * 1, 10) 
        ws_2.column_dimensions[col_letter].width = adjusted_width

    # green for true and red for false
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    header = [cell.value for cell in ws_2[1]]
    try:
        sig_col_idx = header.index('significant') + 1 
    except ValueError:
        sig_col_idx = None
    if sig_col_idx:
        for row in ws_2.iter_rows(min_row=2, max_row=ws_2.max_row):
            cell = row[sig_col_idx - 1] 
            cell_value = str(cell.value).strip() if cell.value is not None else ''
            if cell_value == 'True':
                cell.fill = green_fill
            elif cell_value == 'False':
                cell.fill = red_fill

    wb_2.save(output_analysis_1)

    # Second step of analysis : by linguistic distance

    # New groups for CLE
    def classify_cle(value):
        if value <= 0.25:
            return 1 
        elif value <= 0.50:
            return 2 
        elif value <= 0.75:
            return 3 
        else:
            return 4 

    db['cle_bis'] = db['cle'].apply(classify_cle)

    results_cle = []
    group_colors = {1: 'tab:blue', 2: 'tab:orange', 3: 'tab:green', 4: 'tab:purple'}  
    cle_bis_labels = {
        1: 'cle ≤ 0.25',
        2: '0.25 < cle ≤ 0.50',
        3: '0.50 < cle ≤ 0.75',
        4: 'cle > 0.75'}
    
    # Ploting each CLE per metric
    for metric in metrics:
        plt.figure(figsize=(16, 8))

        for cle_value in sorted(db['cle_bis'].dropna().unique()):
            subset = db[db['cle_bis'] == cle_value]
            daily_avg_cle = subset.groupby('approved_date')[metric].mean().reset_index()
            monthly_avg_cle = daily_avg_cle.groupby(pd.Grouper(key='approved_date', freq='12ME'))[metric].mean().reset_index()

            label_name = cle_bis_labels.get(cle_value, f'CLE group {cle_value}')
            plt.plot(monthly_avg_cle['approved_date'], monthly_avg_cle[metric], 
                    label=label_name, color=group_colors.get(cle_value, None))

            before = subset.loc[subset['approved_date'] <= cutoff_date, metric].dropna()
            after = subset.loc[subset['approved_date'] > cutoff_date, metric].dropna()

            if len(before) > 1 and len(after) > 1:
                t_stat, p_val = ttest_ind(before, after, equal_var=False)
                results_cle.append({
                    'cle_group': cle_value,
                    'metric': metric,
                    't_stat': t_stat,
                    'p_value': p_val,
                    'mean_before': before.mean(),
                    'mean_after': after.mean()
                })
            else:
                results_cle.append({
                    'cle_group': cle_value,
                    'metric': metric,
                    't_stat': None,
                    'p_value': None,
                    'mean_before': before.mean(),
                    'mean_after': after.mean()
                })

        plt.axvline(x=target_date, color='red', linestyle='--', linewidth=1, label=interest_name)
        plt.title(f'Monthly average of {metric.replace("_", " ")} by cle intervals')
        plt.xlabel('Date')
        plt.ylabel('Average value')
        plt.legend(title='cle intervals')
        plt.xticks(ticks, labels, rotation=45)
        plt.grid(True)
        plt.tight_layout()
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        plt.savefig(f'outputs/graphs/by_cle/comparison_monthly_average_{metric}.png')
        plt.close()


    cle_bis_labels = {1: '0-0.25', 2: '0.25-0.50', 3: '0.50-0.75', 4: '0.75-1.00'}
    interest_date = datetime.datetime(interest_year, interest_month, interest_day)
    cutoff_date = pd.Timestamp(interest_date)

    results_intra = []

    # Computing the metrics per CLE
    for metric in metrics[:-1]:
        for group in sorted(db['cle_bis'].dropna().unique()):
            subset = db[db['cle_bis'] == group]
            before = subset[subset['approved_date'] <= cutoff_date][metric].dropna()
            after = subset[subset['approved_date'] > cutoff_date][metric].dropna()

            mean_before = before.mean() if len(before) > 0 else None
            mean_after = after.mean() if len(after) > 0 else None

            taux_var = ((mean_after - mean_before) / abs(mean_before) * 100) if mean_before not in [None, 0] and mean_after is not None else None
            
            # t-test between before and after the event for each metric of each CLE
            if len(before) > 1 and len(after) > 1:
                t_stat, p_val = ttest_ind(before, after, equal_var=False, nan_policy='omit')
            else:
                t_stat, p_val = None, None

            sig = 'True' if p_val is not None and p_val < 0.05 else 'False'

            results_intra.append({
                'metric': metric,
                'cle_group': cle_bis_labels.get(group),
                'mean_before': round(mean_before, 4) if mean_before is not None else None,
                'mean_after': round(mean_after, 4) if mean_after is not None else None,
                't_stat_before_after': round(t_stat, 4) if t_stat is not None else None,
                'p_val_before_after': round(p_val, 4) if p_val is not None else None,
                'significant': sig,
                '%_change': round(taux_var, 2) if taux_var is not None else None
            })
    
    # Computing the global percent in change for each CLE (average of % in change of all metrics)
    df_intra = pd.DataFrame(results_intra)
    df_intra = df_intra.dropna(subset=['%_change'])
    average_abs_change = df_intra.groupby('cle_group')['%_change'].apply(lambda x: round(abs(x).mean(), 2)).reset_index()
    average_abs_change.rename(columns={'%_change': 'avg_abs_%_change'}, inplace=True)

    # Write both sheets in one Excel
    with pd.ExcelWriter(output_analysis_2, engine='openpyxl') as writer:
        df_intra.to_excel(writer, sheet_name='Intra_Group', index=False)
        average_abs_change.to_excel(writer, sheet_name='Avg_Abs_Change', index=False)

    # Format both dataframes
    wb = openpyxl.load_workbook(output_analysis_2)
    for sheet_name in ['Intra_Group', 'Avg_Abs_Change']:
        ws = wb[sheet_name]
        blue_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
        gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.fill = blue_fill
            cell.font = bold_font
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.fill = blue_fill
                cell.font = bold_font

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.row != 1 and cell.column != 1:
                    cell.fill = gray_fill
                cell.border = thin_border

        for i, col in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max(10, max_len * 1.2)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if str(cell.value).strip() == 'True':
                    cell.fill = green_fill
                elif str(cell.value).strip() == 'False':
                    cell.fill = red_fill

    wb.save(output_analysis_2)

    print('Script finished !')


# To run the script not in main

# output_computations = 'outputs/computations.csv'
# output_analysis_1 = 'outputs/analysis_all_papers.xlsx'
# output_analysis_2 = 'outputs/analysis_by_cle.xlsx'
# begin_date = '2015-12-01' # 'YYYY--MM-DD'
# end_date = '2025-07-06'
# interest_name = 'ChatGPT'
# interest_year = 2022
# interest_month = 11
# interest_day = 30

# analysis(output_computations,output_analysis_1,output_analysis_2,interest_name,interest_year,interest_month,interest_day)
